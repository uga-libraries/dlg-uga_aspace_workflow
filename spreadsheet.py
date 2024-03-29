from openpyxl import load_workbook, utils
from openpyxl.styles import PatternFill


class Spreadsheet:
    """Take the spreadsheet of top containers and parse them for aspace and write the spreadsheet template data to the
    template provided by the user"""

    @staticmethod
    def sort_input(container_inputs):
        """
        Intakes a spreadsheet and returns a list of barcodes

        :param str container_inputs: user input values of top container barcodes or URIs

        :return list barcodes: list of barcodes or URIs (str) for top containers
        """
        barcodes = []
        if "," in container_inputs:
            csep_containers = [user_input.strip() for user_input in container_inputs.split(",")]
            for container in csep_containers:
                linebreak_containers = container.splitlines()
                for lb_container in linebreak_containers:
                    barcodes.append(lb_container)
        else:
            barcodes = [user_input.strip() for user_input in container_inputs.splitlines()]
        return barcodes

    @staticmethod
    def get_cell_coordinate(coordinate, row_number):
        """
        Intakes cell coordinate number from spreadsheet and row number to reutn cell coordinate

        :param tuple coordinate: cell coordinate in tuple
        :param int row_number: the row number for each archival object being written

        :return str cell_coordinate: the cell coordinate, ex. A3
        """
        cell_letter = utils.get_column_letter(coordinate[1])
        cell_coordinate = f'{cell_letter}{row_number}'
        return cell_coordinate

    @staticmethod
    def write_template(aspace_dlg_template, arch_obj, resource_obj, row_number, repository):
        """
        Intakes the aspace_dlg spreadsheet and writes data from archival object instance row by row

        :param str aspace_dlg_template: file location for the user input spreadsheet to write to
        :param ArchivalObject arch_obj: instance of ArchivalObject class with data to write to spreadsheet
        :param ResourceObject resource_obj: instance of ResourceObject class with resource data to write to spreadsheet
        :param int row_number: running count of the row number to add data to
        :param str repository: The repository title to fill out in holding institution column
        """
        data_columns = {}
        temp_wb = load_workbook(aspace_dlg_template)
        for sheet in temp_wb:
            for row in sheet.iter_rows(max_row=1, max_col=28):
                for header in row:
                    if header.value is not None:
                        data_columns[header.value] = utils.coordinate_to_tuple(header.coordinate)
            if "Archival Object URI" not in data_columns:
                data_list = list(data_columns.items())
                column_num = int(data_list[-1][1][1]) + 1  # get the last column's number value from converted data_list
                column_letter = utils.get_column_letter(column_num)
                cell_coordinate = f'{column_letter}1'
                sheet[cell_coordinate] = "Archival Object URI"
                data_columns["Archival Object URI"] = utils.coordinate_to_tuple(cell_coordinate)
            for column, coordinate in data_columns.items():
                if "DLG Collection ID" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.dlg_id
                elif "Box, Folder, and Item Number" == column:
                    bfi = [arch_obj.box, arch_obj.child, arch_obj.grandchild]
                    bfi_str = ""
                    for container in bfi:
                        if container:
                            bfi_str += container + ", "
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = bfi_str[:-2]
                elif "record_id" == column:
                    if "?" in arch_obj.record_id:
                        sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.record_id
                        for cell in sheet[f'{row_number}:{row_number}']:
                            cell.fill = PatternFill(start_color='FFFF0000',
                                                    end_color='FFFF0000',
                                                    fill_type='solid')
                    else:
                        sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.record_id
                elif "dcterms_title" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.title
                elif "dcterms_creator" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.creator
                elif "dcterms_subject" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.subject
                elif "dcterms_description" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.description
                elif "dc_date" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.date
                elif "dcterms_spatial" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.subjspa
                elif "dcterms_medium" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.subjmed
                elif "extent" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.extent
                elif "dcterms_language" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.language
                elif "dcterms_bibliographic_citation" == column:
                    if not arch_obj.citation:
                        sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.citation
                    else:
                        sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.citation
                elif "dlg_subject_personal" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = resource_obj.subjper
                elif "holding institution" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = repository
                elif "public" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = "true"
                elif "Archival Object URI" == column:
                    sheet[Spreadsheet.get_cell_coordinate(coordinate, row_number)] = arch_obj.arch_obj_uri
        temp_wb.save(aspace_dlg_template)
